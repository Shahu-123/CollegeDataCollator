import requests
import openai
from openai import OpenAI
from openpyxl import Workbook, load_workbook

# Define the OpenAI and Google API keys
client = OpenAI(api_key="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")
google_api_key = 'AIzaSyCQJGwuvHtTC4bwDyZzc8zlhKyHMRW3F4E'
cx = '329b5b7a4d268471c'
openai_api_key = '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'
filename = "/Users/shahu/Desktop/CollegeList.xlsx"
# Initialize OpenAI client
openai.api_key = openai_api_key

def get_search_results(query, api_key, cx):
    url = f"https://www.googleapis.com/customsearch/v1?q={query}&key={api_key}&cx={cx}"
    response = requests.get(url)
    try:
        response.raise_for_status()  # Raises a HTTPError for bad responses
        return response.json()  # Returns the search results as a JSON object
    except requests.exceptions.HTTPError as e:
        return f"HTTP error occurred: {e}"
    except Exception as e:
        return f"Other error occurred: {e}"

def print_snippets(search_results):
    text = ""
    if 'items' in search_results:
        for item in search_results['items']:
            text += item['snippet'] + " "
    else:
        print("No results found")
    return text.strip()

def read_cells_from_excel(file_path, sheet_name, cell_range='A1:A52'):
    # Load the workbook and select the specified sheet
    workbook = load_workbook(filename=file_path)
    sheet = workbook[sheet_name]

    # Extract the values from the specified range and store them in a list
    cell_values = []
    for row in sheet[cell_range]:
        for cell in row:
            cell_values.append(str(cell.value) if cell.value is not None else "")

    return cell_values

def extract_specific_info_from_chatgpt(text, required_info, api_key):
    try:
        chat_completion = client.chat.completions.create(
            messages=[
                {
                    "role": "user",
                    "content": f"Extract the exact answer representing the {required_info} for undergraduates from the following text: {text}. Output the answer and ONLY the answer, otherwise if you are not confident, do not output anything.",
                }
            ],
            model="gpt-3.5-turbo",
            temperature=0
        )
        return chat_completion.choices[0].message.content.strip()
    except Exception as e:
        return f"An error occurred: {e}"


def write_to_excel(data_dict, file_path, sheet_name):
    # Load the workbook and select the specified sheet
    workbook = load_workbook(filename=file_path)
    sheet = workbook[sheet_name]

    # Iterate through each row in column A to find matches with dictionary keys
    for row in sheet.iter_rows(min_col=1, max_col=1, min_row=1, max_row=sheet.max_row):
        for cell in row:
            key = cell.value
            if key in data_dict:
                # Get the corresponding data list for this key
                data = data_dict[key]

                # Determine the starting and ending columns for insertion (C to P -> 3 to 16)
                start_col = 3
                end_col = start_col + len(data) - 1

                # Insert the data into the appropriate cells
                for col, value in enumerate(data, start=start_col):
                    sheet.cell(row=cell.row, column=col, value=value)

    # Save the updated workbook
    workbook.save(filename=file_path)

details = read_cells_from_excel(filename, sheet_name="College List Matrix")
colleges = ["MIT", "Carnegie Mellon", "Stanford", "UC Berkeley", "Harvard", "Princeton", "Cornell", "UCLA", "Columbia", "NYU", "Georgia Tech", "Yale", "UPenn", "Caltech"]
results = {}

row_count = 0
for detail in details:
    row_count += 1
    check = input(f"Enter 'n' to skip detail for {detail}:\n")
    if check == 'n':
        continue
    record = []
    for college in colleges:
        query = f"{college}'s {detail}"
        search_results = get_search_results(query, google_api_key, cx)
        info = print_snippets(search_results)
        extracted_info = extract_specific_info_from_chatgpt(info, detail, openai_api_key)
        record.append(extracted_info)
        print(f"The {detail} for {college} is {extracted_info}")
    results[detail] = record

write_to_excel(results, filename, sheet_name="College List Matrix")
